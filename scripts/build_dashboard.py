"""生成模拟盘 Dashboard 数据文件 (data.js)
用法:
  python3 build_dashboard.py <模拟盘Excel文件路径>

示例:
  python3 build_dashboard.py 模拟盘_V46_v62_v20_final_20260624_134632.xlsx

依赖: openpyxl (pip install openpyxl)
"""
import sys, json, os
from collections import defaultdict
from datetime import datetime, timedelta

def build(excel_path, pool_path='股票池.xlsx'):
    """读取模拟盘Excel, 生成 data.js"""
    
    import openpyxl
    
    # 1. 读取股票池行业
    code_to_industry = {}
    if os.path.exists(pool_path):
        wb = openpyxl.load_workbook(pool_path, data_only=True, read_only=True)
        for row in wb['股票池'].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                code = str(row[0]).strip()
                ind = str(row[2]).strip() if row[2] else ''
                if code and ind:
                    code_to_industry[code] = ind
        wb.close()
    else:
        print(f'⚠️  未找到股票池文件: {pool_path}, 行业信息将显示为"未知"')
    
    # 2. 读取模拟盘Excel
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    
    # 统计信息
    ws6 = wb['统计分析']
    stats = {}
    for row in ws6.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None:
            stats[str(row[0])] = row[1]
    
    # 净值曲线
    ws3 = wb['净值曲线']
    nav_data = []
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if row[0]:
            nav_data.append({'date': str(row[0]), 'nav': row[1], 'positions': row[2]})
    
    # 当前持仓
    ws1 = wb['当前持仓']
    h1 = [c.value for c in next(ws1.iter_rows(min_row=1, max_row=1))]
    holdings = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row[0]:
            h = dict(zip(h1, [str(x) if x is not None else '' for x in row]))
            h['行业'] = code_to_industry.get(h['代码'], '未知')
            h['指数归属'] = h.get('指数归属', '其他')
            h['池内'] = h.get('池内', '是')
            holdings.append(h)
    
    # 行业持仓分布
    ind_val = defaultdict(float)
    for h in holdings:
        ind_val[h['行业']] += float(h['市值']) if h.get('市值') else 0
    industry_data = sorted(
        [{'industry': k, 'value': round(v, 2)} for k, v in ind_val.items()],
        key=lambda x: -x['value']
    )
    
    # 全部交易
    ws2 = wb['交易记录']
    h2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    all_trades = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            t = dict(zip(h2, [str(x) if x is not None else '' for x in row]))
            t['行业'] = code_to_industry.get(t['代码'], '未知')
            all_trades.append(t)
    
    wb.close()
    
    # 最近一周交易
    latest = max(t['日期'] for t in all_trades)
    ld = datetime.strptime(latest, '%Y-%m-%d')
    week_ago = ld - timedelta(days=7)
    week_trades = [t for t in all_trades if datetime.strptime(t['日期'], '%Y-%m-%d') >= week_ago]
    
    # 卖出交易分析
    sell_trades = [t for t in all_trades if t['买卖方向'] == '卖出']
    for t in sell_trades:
        t['盈亏%值'] = float(t['盈亏%']) if t.get('盈亏%') else 0.0
    sell_sorted = sorted(sell_trades, key=lambda x: -x['盈亏%值'])
    top_winners = sell_sorted[:10]
    top_losers = sell_sorted[-10:][::-1] if len(sell_sorted) >= 10 else sell_sorted[::-1]
    
    # 行业盈亏
    ind_pnl = defaultdict(lambda: {'cnt': 0, 'wins': 0, 'sum_pnl': 0.0})
    for t in sell_trades:
        ind = t.get('行业', '未知')
        ind_pnl[ind]['cnt'] += 1
        ind_pnl[ind]['sum_pnl'] += t['盈亏%值']
        if t['盈亏%值'] > 0:
            ind_pnl[ind]['wins'] += 1
    ind_pnl_list = sorted(
        [{'industry': k, 'avg_pnl': round(v['sum_pnl']/v['cnt'], 2) if v['cnt'] else 0,
          'win_rate': round(v['wins']/v['cnt']*100, 1) if v['cnt'] else 0}
         for k, v in ind_pnl.items()],
        key=lambda x: -x['avg_pnl']
    )[:12]
    
    # 卖出原因
    sell_reasons = {}
    for k, v in stats.items():
        if '卖出原因' in str(k):
            sell_reasons[str(k).replace('卖出原因:', '')] = v
    
    # 盈亏分布
    bins = {'<-10%': 0, '-10~-5%': 0, '-5~0%': 0, '0~5%': 0, '5~10%': 0, '>10%': 0}
    for t in sell_trades:
        p = t['盈亏%值']
        if p < -10: bins['<-10%'] += 1
        elif p < -5: bins['-10~-5%'] += 1
        elif p < 0: bins['-5~0%'] += 1
        elif p < 5: bins['0~5%'] += 1
        elif p < 10: bins['5~10%'] += 1
        else: bins['>10%'] += 1
    
    # 组装输出
    output = {
        'stats': stats, 'nav': nav_data,
        'holdings': holdings, 'industry': industry_data,
        'week_trades': week_trades,
        'top_winners': top_winners, 'top_losers': top_losers,
        'ind_pnl': ind_pnl_list, 'sell_reasons': sell_reasons,
        'pnl_dist': [{'r': k, 'c': v} for k, v in sorted(bins.items())],
        'updated': datetime.now().strftime('%Y-%m-%d %H:%M')
    }
    
    # 写入 data.js
    out_name = 'data.js'
    if len(sys.argv) > 2:
        out_name = sys.argv[2]
    js_path = os.path.join(os.path.dirname(excel_path) or '.', out_name)
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write('const D = ')
        json.dump(output, f, ensure_ascii=False, indent=2)
        f.write(';\n')
    
    print(f'✅ data.js 已生成: {os.path.abspath(js_path)}')
    print(f'   净值: {len(nav_data)}天, 持仓: {len(holdings)}只, 交易: {len(all_trades)}条')
    print(f'   最近一周交易: {len(week_trades)}条')
    print(f'\n📌 将 data.js 和 V46_Dashboard_20260624.html 放在同一目录,')
    print(f'   双击 HTML 文件即可查看 Dashboard')

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f'❌ 文件不存在: {excel_path}')
        sys.exit(1)
    build(excel_path)
