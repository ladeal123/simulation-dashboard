"""V48 v29 + 波动率惩罚 (恢复因子+时间衰减)
- 第一排序: BAR% 10日斜率 → Top 100
- 第二排序: 恢复因子(时间衰减后) - 0.15×vol20
- 买入: 第二排序结果的前 N 只
- 仅涨停过滤, 无5%涨幅限制
- 回测: 2020-06-01 ~ 2026-05-14
"""
import csv, math, sys, json, datetime
# V42 防御板块差异化止盈
import openpyxl as _openpyxl_w42
_wb_w42 = _openpyxl_w42.load_workbook('/workspace/股票池_new.xlsx', data_only=True)
_ws_w42 = _wb_w42['股票池']
_code_to_industry_w42 = {}
for _r in range(2, _ws_w42.max_row + 1):
    _c = _ws_w42.cell(row=_r, column=1).value
    _i = _ws_w42.cell(row=_r, column=3).value
    if _c and _i:
        _code_to_industry_w42[str(_c)] = str(_i)
DEFENSIVE_INDUSTRIES = {'银行', '房地产', '钢铁', '公用事业', '交通运输', '煤炭', '石油石化', '建筑材料', '建筑装饰', '纺织服饰', '美容护理'}
print(f'[V42] 防御板块行业数: {len(DEFENSIVE_INDUSTRIES)}, 加载完成: {len(_code_to_industry_w42)} 只', flush=True)
from collections import defaultdict, Counter

# ===== 参数 =====
FAST, SLOW, SIGNAL = 40, 80, 32
START_DATE = '2025-07-01'
END_DATE = '2026-06-22'
INIT_CASH = 10_000_000
POS_SIZE = 0.02
MAX_POS = 50
MIN_HOLD = 3
MIN_HOLD_RANK = 3
STOP_LOSS = -0.25
SLOPE_WIN = 10  # V48 v3: 10日斜率
VOL_WIN = 10  # V29_vol20: 20日波动率窗口
VOL_ALPHA = 0.15  # V29_vol10: 波动率惩罚系数 α=0.15_win10
DIF_BONUS = 0.1    # DIF%额外权重 (实验B)
POOL_TOP_N = 200
POOL_SIZE = 150  # 最终版: 第一排序候选池大小
DIF_PCT_FLOOR = 0.0  # V48 v4: DIF% > 0 (vs v3 的 > -3%)
BAR_FLOOR = 0.0       # BAR > 0
# V48 v12: V46 风格的仓位管理 + 平滑调整
MAX_POS_BULL = 50
MAX_POS_FLAT = 20
MAX_POS_SHORT = 0  # V48 v30=v12 v3 终版: short 清仓
STATE_TARGET = {'bull': 50, 'flat': 30, 'short': 10}
SMOOTH_ALPHA = 0.7
INDEX_FILE = '/workspace/user_input_files/中证1000指数价格后复权.xlsx'
INDEX_FAST = 10; INDEX_SLOW = 20; INDEX_DEA = 9  # V4 fast MACD
TH_SHORT = 0.0; BOTTOM = -3.0  # V46 真实参数 (th_short=0.0)

# ===== 涨停判断 =====
def is_limit_up_open(code, price, prev_close):
    """判断开盘是否涨停。主板≥9.5%, 科创/创业板≥19.5%"""
    if prev_close <= 0 or price <= 0: return False
    chg = (price - prev_close) / prev_close
    if code.startswith('688') or code.startswith('300') or code.startswith('301'):
        return chg >= 0.195
    return chg >= 0.095

# ===== EMA (V46 主回测风格: 价格初始化) =====
def ema_v46(values, n_):
    """价格初始化: 跳过 0/None, 第一个有效值=价格本身"""
    a = 2.0/(n_+1); out = [None]*len(values)
    fv = None
    for i, v in enumerate(values):
        if v is not None and v > 0:
            fv = i; break
    if fv is None: return out
    out[fv] = values[fv]
    for i in range(fv+1, len(values)):
        v = values[i]
        if v is None or v == 0:
            out[i] = out[i-1]
        else:
            out[i] = a*v + (1-a)*out[i-1]
    return out

def calc_slope(values, win=5):
    out = [None]*len(values)
    for i in range(win-1, len(values)):
        y = [values[i-j] for j in range(win-1, -1, -1)]
        if any(v is None for v in y): continue
        xm = [j - (win-1)/2 for j in range(win)]
        ym = [yi - sum(y)/win for yi in y]
        num = sum(xm[j]*ym[j] for j in range(win))
        den = sum(xm[j]**2 for j in range(win))
        out[i] = num/den if den != 0 else 0.0
    return out

def dif_slope(dif, win=2):
    """V46 风格: 2日 OLS 回归斜率"""
    out=[0.0]*len(dif)
    for i in range(win,len(dif)):
        y=[dif[i-j] for j in range(win-1,-1,-1)]
        xm=[j-(win-1)/2 for j in range(win)]
        ym=[yi-sum(y)/win for yi in y]
        num=sum(xm[j]*ym[j] for j in range(win))
        den=sum(xm[j]**2 for j in range(win))
        out[i]=num/den if den!=0 else 0
    return out

def calc_v4_state_v46(idx_dates, prices, dif, dif_pct, dea, slope, bar=None, th_short=0.0, bottom=-3.0):
    """V46 v4fix_v3 修复版: 提前入失败退出
    修复 1: gcd 死叉永远检测 (无 BAR 屏蔽)
    修复 2: 提前入失败退出 (前日 BAR<0 + 当日反弹失败 → pos=0)
    返回: list of state, len = len(dates)-1
    state: 1=bull, 0=flat, -1=short
    """
    pos = 0
    result = []
    prev_bar = None
    for i in range(1, len(idx_dates)):
        rd = dif[i] if dif[i] is not None else 0
        rdea = dea[i] if dea[i] is not None else 0
        pd_ = dif[i-1] if dif[i-1] is not None else 0
        pdea = dea[i-1] if dea[i-1] is not None else 0
        rdp = dif_pct[i]
        rsl = slope[i]
        rbar = bar[i] if (bar is not None and i < len(bar) and bar[i] is not None) else 0
        gcu = pd_ <= pdea and rd > rdea
        gcd = pd_ >= pdea and rd < pdea
        if pos == 0:
            if gcu: pos = 1
            elif rdp > 0 and rsl > 0.02: pos = 1
            elif rdp < bottom and rsl > 0.03: pos = 1
            elif rdp <= th_short: pos = -1
        elif pos == 1:
            # 修复 1: gcd 死叉永远检测, 不受 BAR 影响
            if gcd:
                pos = 0
                if pos == 0 and rdp <= th_short: pos = -1
            # 修复 2: 提前入失败退出
            elif prev_bar is not None and prev_bar < 0:
                if rdp > 0 and rsl < -0.02:
                    pos = 0
        elif pos == -1:
            if gcu: pos = 1
            elif rdp < bottom and rsl > 0.03: pos = 1
        result.append(pos)
        prev_bar = rbar
    return result

def calc_v4_state(dates, dif_pct, dea, th_short=2.0, bottom=-3.0):
    """V46 主回测 calc_pos10 风格的 V4 状态机
    返回: list of (state, dif_pct) for each date, len = len(dates)-1
    state: 1=bull, 0=flat, -1=short
    """
    result = []
    pos = 0
    for i in range(1, len(dates)):
        rd_pct = dif_pct[i] if dif_pct[i] is not None else 0.0
        rd = rd_pct  # 用 dif_pct 替代原 raw DIF
        rdea = dea[i] if dea[i] is not None else 0.0
        pd_pct = dif_pct[i-1] if dif_pct[i-1] is not None else 0.0
        pdea = dea[i-1] if dea[i-1] is not None else 0.0
        gcu = pd_pct <= pdea and rd_pct > rdea
        gcd = pd_pct >= pdea and rd_pct < rdea
        # 简化 slope = (rd_pct - pd_pct)
        slope_proxy = rd_pct - pd_pct
        if pos == 0:
            if gcu: pos = 1
            elif rd_pct > 0 and slope_proxy > 0.02: pos = 1
            elif rd_pct < bottom and slope_proxy > 0.03: pos = 1
            elif rd_pct <= -th_short: pos = -1
        elif pos == 1:
            if gcd:
                if rd_pct <= -th_short: pos = -1
                else: pos = 0
        elif pos == -1:
            if gcu: pos = 1
            elif rd_pct < bottom and slope_proxy > 0.03: pos = 1
        result.append((pos, rd_pct))
    return result
    """5点线性回归斜率 (OLS)"""
    out = [None]*len(values)
    for i in range(win-1, len(values)):
        y = [values[i-j] for j in range(win-1, -1, -1)]
        if any(v is None for v in y): continue
        xm = [j - (win-1)/2 for j in range(win)]
        ym = [yi - sum(y)/win for yi in y]
        num = sum(xm[j]*ym[j] for j in range(win))
        den = sum(xm[j]**2 for j in range(win))
        out[i] = num/den if den != 0 else 0.0
    return out

# ===== 加载数据 (模拟盘xlsx) =====
print("加载数据(模拟盘xlsx)...")
sys.stdout.flush()
import openpyxl as _xl

# 改用模拟盘的xlsx
_wb_data = _xl.load_workbook('/workspace/模拟盘收盘价后复权.xlsx', data_only=True, read_only=True)
_ws_data = _wb_data['close_price']
_data_header = list(_ws_data.iter_rows(min_row=1, max_row=1, values_only=True))[0]
_data_codes = [str(c).strip() for c in _data_header[1:] if c is not None]

stock_data = defaultdict(dict)
for _row in _ws_data.iter_rows(min_row=5, values_only=True):
    _d = str(_row[0])[:10] if _row[0] else ''
    if not _d.startswith('202'): continue
    for _i, _code in enumerate(_data_codes):
        if _i+1 >= len(_row): break
        _v = _row[_i+1]
        if _v is not None and isinstance(_v, (int, float)) and _v > 0:
            stock_data[_code][_d] = _v
_wb_data.close()

all_dates = sorted({d for code in stock_data for d in stock_data[code]})
sim_dates = [d for d in all_dates if START_DATE <= d <= END_DATE]
print(f"股票数: {len(stock_data)}, 总交易日: {len(all_dates)}")
print(f"回测区间: {sim_dates[0]} ~ {sim_dates[-1]} ({len(sim_dates)} 天)")

# ===== 加载指数数据 + 计算 V4 状态 =====
print("加载中证1000指数并计算V4状态...")
sys.stdout.flush()
import pandas as pd
idx_df = pd.read_excel(INDEX_FILE, header=None, skiprows=3)
idx_df.columns = ['date', 'price']
idx_data = {}
for _, row in idx_df.iterrows():
    try:
        d = str(row['date'])[:10]
        p = float(row['price'])
        if p > 0: idx_data[d] = p
    except: continue
idx_dates = sorted(idx_data.keys())
idx_ef = ema_v46([idx_data[d] for d in idx_dates], INDEX_FAST)
idx_es = ema_v46([idx_data[d] for d in idx_dates], INDEX_SLOW)
idx_dif = [idx_ef[i] - idx_es[i] if (idx_ef[i] is not None and idx_es[i] is not None) else None for i in range(len(idx_dates))]
idx_dif_pct = [idx_dif[i]/idx_data[idx_dates[i]]*100 if (idx_dif[i] is not None and idx_data[idx_dates[i]] > 0) else 0.0 for i in range(len(idx_dates))]
idx_dea = ema_v46(idx_dif, INDEX_DEA)
idx_bar = [(idx_dif[i] - idx_dea[i]) * 2 if (idx_dif[i] is not None and idx_dea[i] is not None) else None for i in range(len(idx_dates))]
idx_slope = dif_slope(idx_dif_pct, 2)  # V46 风格: 2日 OLS 斜率
v4_states = calc_v4_state_v46(idx_dates, [idx_data[d] for d in idx_dates], idx_dif, idx_dif_pct, idx_dea, idx_slope, idx_bar, TH_SHORT, BOTTOM)
# 映射: 日期 -> state
idx_state_map = {idx_dates[i+1]: v4_states[i] for i in range(len(v4_states))}
print(f"指数日期: {len(idx_dates)}, V4状态计算完成 (V46真实逻辑)")

# ===== 预计算指标 =====
print("预计算MACD(40/80/32)和5日斜率...")
sys.stdout.flush()
IND = {}
for code, dp in stock_data.items():
    d_list = sorted(dp.keys())
    if len(d_list) < 200: continue  # MACD(40/80/32) + 5日斜率 + 安全余量
    prices = [dp.get(d, 0) for d in d_list]
    ef = ema_v46(prices, FAST)
    es = ema_v46(prices, SLOW)
    dif = [ef[i] - es[i] if (ef[i] is not None and es[i] is not None) else None for i in range(len(prices))]
    dea = ema_v46(dif, SIGNAL)
    bar = [(dif[i] - dea[i]) * 2 if (dif[i] is not None and dea[i] is not None) else None for i in range(len(dif))]
    dif_pct = [dif[i]/prices[i]*100 if (dif[i] is not None and prices[i] > 0) else None for i in range(len(dif))]
    bar_pct = [bar[i]/prices[i]*100 if (bar[i] is not None and prices[i] > 0) else None for i in range(len(bar))]
    bar_pct_slope = calc_slope(bar_pct, SLOPE_WIN)
    date_to_idx = {d: i for i, d in enumerate(d_list)}
    IND[code] = {
        'dates': d_list, 'date_to_idx': date_to_idx,
        'prices': prices, 'ef': ef, 'es': es, 'dif': dif, 'dea': dea, 'bar': bar,
        'dif_pct': dif_pct, 'bar_pct': bar_pct, 'bar_pct_slope': bar_pct_slope
    }
    # V29_vol10: 计算近 20 日日波动率 (std of daily returns)
    vol20_list = [None] * len(prices)
    for j in range(1, len(prices)):
        if j < VOL_WIN: continue
        window = prices[j-VOL_WIN+1:j+1]
        if any(p is None or p <= 0 for p in window): continue
        rets = [(window[k] - window[k-1]) / window[k-1] for k in range(1, len(window))]
        if len(rets) >= 2:
            mean_r = sum(rets) / len(rets)
            var_r = sum((r - mean_r)**2 for r in rets) / (len(rets) - 1)
            vol20_list[j] = (var_r ** 0.5) * 100  # 百分比形式
    IND[code]['vol20'] = vol20_list
    # BAR回调再扩张因子: 今日BAR / 近10日最小正BAR, 加时间衰减惩罚
    RECOVERY_WIN = 10
    # 时间衰减系数β可从命令行传入: python3 xxx.py [beta]
    TIME_DECAY_BETA = 0.20  # 时间衰减系数 (固定最优值)
    bar_recovery = [None] * len(prices)
    for j in range(len(prices)):
        if bar_pct[j] is None or bar_pct[j] <= 0:
            bar_recovery[j] = None
            continue
        if j < RECOVERY_WIN: continue
        # 找到近10日内最小正BAR及其位置
        min_val = None
        min_idx = None
        for k in range(j-RECOVERY_WIN+1, j+1):
            if bar_pct[k] is not None and bar_pct[k] > 0:
                if min_val is None or bar_pct[k] < min_val:
                    min_val = bar_pct[k]
                    min_idx = k
        if min_val is None: continue
        # 统计正BAR数量
        n_pos = sum(1 for k in range(j-RECOVERY_WIN+1, j+1) if bar_pct[k] is not None and bar_pct[k] > 0)
        if n_pos < 3: continue  # 至少3个正BAR
        raw_recovery = bar_pct[j] / (min_val + 1e-8)
        # 时间衰减: 距离最低点越久, 因子打折越多
        days_since_min = j - min_idx  # 最小正BAR是几天前
        time_penalty = 1 + TIME_DECAY_BETA * max(0, days_since_min - 1)  # 昨天最低不打折
        bar_recovery[j] = raw_recovery / time_penalty
    IND[code]['bar_recovery'] = bar_recovery
    # V46 短周期MACD(10,20,9) 用于死叉卖出信号
    V46_F, V46_S, V46_SIG = 10, 20, 9
    v46_ef = ema_v46(prices, V46_F)
    v46_es = ema_v46(prices, V46_S)
    v46_dif = [v46_ef[i]-v46_es[i] if (v46_ef[i] is not None and v46_es[i] is not None) else None for i in range(len(prices))]
    v46_dea = ema_v46(v46_dif, V46_SIG)
    IND[code]['v46_dif'] = v46_dif
    IND[code]['v46_dea'] = v46_dea
print(f"有效股票: {len(IND)}")

# ===== 加载成交量数据 =====
print("加载成交量数据...")
import pandas as pd
_df_vol = pd.read_excel('/workspace/股票池成交量.xlsx', engine='calamine', header=None)
_vol_codes = [str(c).strip() for c in _df_vol.iloc[0, 1:].tolist() if c is not None]
vol_data = {}
for _ri in range(7, len(_df_vol)):
    _row = _df_vol.iloc[_ri]; _v0 = _row[0]
    _d = _v0.strftime('%Y-%m-%d') if isinstance(_v0, datetime.datetime) else str(_v0)[:10]
    if not _d.startswith('20'): continue
    for _j, _code in enumerate(_vol_codes):
        if _j + 1 >= len(_row): break
        _v = _row[_j + 1]
        if _v is not None and isinstance(_v, (int, float)) and _v > 0:
            vol_data.setdefault(_code, {})[_d] = float(_v)
del _df_vol
print(f"  成交量: {len(vol_data)}只有数据")

def get_vol_ratio(code, date):
    """量比 = 当日量 / 近5日均量"""
    vd = vol_data.get(code)
    if not vd: return None
    v_today = vd.get(date)
    if not v_today or v_today <= 0: return None
    try: idx = all_dates.index(date)
    except: return None
    v5_sum = 0; v5_cnt = 0
    for i in range(max(0, idx-4), idx):
        v = vd.get(all_dates[i])
        if v and v > 0: v5_sum += v; v5_cnt += 1
    if v5_cnt < 3: return None
    return v_today / (v5_sum / v5_cnt)

# ===== 加载开盘价(模拟盘xlsx) =====
print("加载开盘价(模拟盘xlsx)...")
import pandas as pd
_xl_wb = _xl.load_workbook('/workspace/模拟盘开盘价后复权.xlsx', data_only=True, read_only=True)
_xl_ws = _xl_wb['close_price']
_xl_rows = list(_xl_ws.iter_rows(values_only=True))
_xl_wb.close()
_open_codes = [str(c).strip() for c in _xl_rows[0][1:] if c is not None]
_open_data = {}
for _r in _xl_rows[4:]:
    if not _r or _r[0] is None: continue
    _d = str(_r[0])[:10] if not hasattr(_r[0], 'strftime') else _r[0].strftime('%Y-%m-%d')
    if not _d.startswith('20'): continue
    for _i, _code in enumerate(_open_codes):
        if _i+1 >= len(_r): break
        _v = _r[_i+1]
        if _v is not None and isinstance(_v, (int,float)) and _v > 0:
            if _code not in _open_data: _open_data[_code] = {}
            _open_data[_code][_d] = float(_v)
print(f"  开盘价: {len(_open_codes)}只, {sum(len(v) for v in _open_data.values())//max(len(_open_codes),1)}天")

def get_open(code, date):
    d = _open_data.get(code)
    return d.get(date) if d else None

# ===== 主回测循环 (开盘价出信号 + 开盘价执行) =====
print("\n开始回测 (开盘价出信号, 开盘价执行)...")
sys.stdout.flush()
portfolio = {}  # code -> {shares, entry_price, entry_date, entry_di, score}
cash = INIT_CASH
trades = []
daily_nav = []
trade_id = 0
n = len(sim_dates)
smooth_max_pos = float(MAX_POS_BULL)  # 平滑仓位值

# MACD EMA系数 (一次性计算)
ALPHA_FAST = 2.0 / (FAST + 1)
ALPHA_SLOW = 2.0 / (SLOW + 1)
ALPHA_DEA = 2.0 / (SIGNAL + 1)

for di, date in enumerate(sim_dates):
    # 1. 构建候选池 (前日用收盘价MACD, 当日用开盘价重算)
    candidates = []
    for code, ind in IND.items():
        i = ind['date_to_idx'].get(date)
        if i is None or i < 1: continue
        open_p = get_open(code, date)
        if open_p is None or open_p <= 0: continue
        
        # ★★★ 用开盘价重算今日MACD ★★★
        ef_y = ind['ef'][i-1]   # 昨天fast EMA
        es_y = ind['es'][i-1]   # 昨天slow EMA
        if ef_y is None or es_y is None: continue
        ef_t = ALPHA_FAST * open_p + (1 - ALPHA_FAST) * ef_y
        es_t = ALPHA_SLOW * open_p + (1 - ALPHA_SLOW) * es_y
        dif = ef_t - es_t
        
        dea_y = ind['dea'][i-1]  # 昨天DEA
        if dea_y is None: continue
        dea_t = ALPHA_DEA * dif + (1 - ALPHA_DEA) * dea_y
        bar = (dif - dea_t) * 2
        
        # BAR%斜率用预计算的(基于收盘, 用昨天数据)
        slope = ind['bar_pct_slope'][i-1]
        if slope is None: continue
        if bar <= BAR_FLOOR: continue
        dif_pct_v = dif / open_p * 100
        if dif_pct_v <= DIF_PCT_FLOOR: continue
        
        # BAR回调再扩张因子 (用昨天数据, 避免先见之明)
        br = ind['bar_recovery'][i-1] if i-1 < len(ind['bar_recovery']) else None
        if br is None: continue
        
        candidates.append((slope, code, open_p, bar/open_p*100, dif_pct_v, slope, br))

    # 第一排序: BAR% 10日斜率 desc → Top 100
    candidates.sort(key=lambda x: -x[0])
    top_pool = candidates[:POOL_SIZE]
    # 第二排序: BAR回调再扩张因子 - α × vol20 (用昨天数据)
    _i_vol_base = (IND.get(top_pool[0][1], {}).get('date_to_idx', {}).get(date)) if top_pool else None
    _i_vol = _i_vol_base - 1 if (_i_vol_base is not None and _i_vol_base > 0) else None
    def _sort2_score(c):
        s = c[6] if c[6] is not None else -1e9  # BAR回调再扩张因子
        dif_pct = c[4] if c[4] is not None else 0  # DIF%
        code_c = c[1]
        ind_c = IND.get(code_c)
        if ind_c is None or _i_vol is None: return s
        vol_arr = ind_c.get('vol20')
        if vol_arr is None or _i_vol >= len(vol_arr): return s
        v = vol_arr[_i_vol] if vol_arr[_i_vol] is not None else 0
        # 实验B: DIF%加入第二排序(水上加分,水下减分)
        return s - VOL_ALPHA * v + DIF_BONUS * dif_pct
    top_pool_sorted = sorted(top_pool, key=lambda x: -_sort2_score(x))
    # code -> 在第二排序中的位置
    code_to_rank = {c[1]: idx+1 for idx, c in enumerate(top_pool_sorted)}

    # 2. SELL: 止损/排名跌出 (用开盘价判断)
    to_sell = []
    for code, pos in list(portfolio.items()):
        ind = IND[code]
        i = ind['date_to_idx'].get(date)
        if i is None: continue
        cur_price = get_open(code, date)  # 用开盘价判断是否卖出
        if cur_price is None or cur_price <= 0: continue
        hold_days = di - pos['entry_di']
        pnl_pct = (cur_price - pos['entry_price']) / pos['entry_price']
        rank = code_to_rank.get(code, 9999)
        is_defensive = _code_to_industry_w42.get(code, '') in DEFENSIVE_INDUSTRIES
        # max_price 用昨日收盘价追踪 (开盘时知道的是昨日收盘, 不是今日)
        if i > 0:
            prev_close = ind['prices'][i-1]
            if prev_close is not None and prev_close > 0:
                if 'max_price' not in pos:
                    pos['max_price'] = prev_close
                if prev_close > pos['max_price']:
                    pos['max_price'] = prev_close
        max_dd_in_hold = (pos['max_price'] - cur_price) / pos['max_price'] if pos['max_price'] > 0 else 0
        pnl_high = (pos['max_price'] - pos['entry_price']) / pos['entry_price']

        if is_defensive:
            if hold_days >= MIN_HOLD and pnl_pct <= -0.15:
                to_sell.append((code, cur_price, pnl_pct, '防御止损', hold_days))
            elif max_dd_in_hold >= 0.05 and pnl_high > 0 and hold_days >= 3:
                to_sell.append((code, cur_price, pnl_pct, '防御移动止盈', hold_days))
            elif hold_days >= 5 and pnl_pct < 0.10:
                to_sell.append((code, cur_price, pnl_pct, '防御超时', hold_days))
            elif rank > 100 and hold_days >= MIN_HOLD_RANK:
                to_sell.append((code, cur_price, pnl_pct, '防御排名跌出100', hold_days))
        else:
            if hold_days >= MIN_HOLD and pnl_pct <= STOP_LOSS:
                to_sell.append((code, cur_price, pnl_pct, '止损', hold_days))
            elif rank > POOL_TOP_N and hold_days >= MIN_HOLD_RANK:
                to_sell.append((code, cur_price, pnl_pct, '排名跌出200', hold_days))
        # V46短周期死叉+量比确认卖出
        if hold_days >= MIN_HOLD:
            ind_v = IND.get(code)
            if ind_v:
                i_v = ind_v['date_to_idx'].get(date)
                if i_v is not None and i_v >= 1:
                    d46 = ind_v['v46_dif'][i_v]; e46 = ind_v['v46_dea'][i_v]
                    dp46 = ind_v['v46_dif'][i_v-1]; ep46 = ind_v['v46_dea'][i_v-1]
                    if not any(x is None for x in [d46, e46, dp46, ep46]):
                        if dp46 > ep46 and d46 <= e46:  # 死叉
                            vol_r = get_vol_ratio(code, date)
                            if True:
                                to_sell.append((code, cur_price, pnl_pct, 'V46死叉+量比', hold_days))

    # 2.5 V48 v12: 仓位管理 (V46 风格 bull/flat/short) + 平滑调整
    # ★★★ T日开盘用T-1日收盘的V4状态(开盘时不知道今日收盘) ★★★
    prev_idx_date = sim_dates[di-1] if di > 0 else date
    idx_state_raw = idx_state_map.get(prev_idx_date, 1)  # 1=bull, 0=flat, -1=short
    state_key = 'bull' if idx_state_raw == 1 else ('short' if idx_state_raw == -1 else 'flat')
    target_max_pos = STATE_TARGET.get(state_key, MAX_POS_BULL)
    smooth_max_pos += SMOOTH_ALPHA * (target_max_pos - smooth_max_pos)
    max_pos = max(5, min(50, int(round(smooth_max_pos))))
    # short 时清仓 (用开盘价判断)
    if max_pos == 0 and len(portfolio) > 0:
        for code, pos in list(portfolio.items()):
            cur_price = get_open(code, date)
            if cur_price is None or cur_price <= 0: continue
            to_sell.append((code, cur_price, (cur_price-pos['entry_price'])/pos['entry_price'], 'short清仓', di-pos['entry_di']))
    # 超过上限时强制卖出（用开盘价判断）
    elif len(portfolio) > max_pos:
        port_with_score = []
        for code in list(portfolio.keys()):
            rank = code_to_rank.get(code, 9999)
            port_with_score.append((rank, code))
        port_with_score.sort(key=lambda x: -x[0])
        for rank, code in port_with_score[:len(portfolio)-max_pos]:
            cur_price = get_open(code, date)
            if cur_price is None or cur_price <= 0: continue
            pnl = (cur_price - portfolio[code]['entry_price']) / portfolio[code]['entry_price']
            to_sell.append((code, cur_price, pnl, f'超仓位({max_pos})', di - portfolio[code]['entry_di']))

    # 执行卖出 (用开盘价执行)
    sold_set = set()
    for code, p_signal, pnl_pct_signal, reason, hd in to_sell:
        if code not in portfolio or code in sold_set: continue
        sold_set.add(code)
        exec_p = get_open(code, date)
        if exec_p is None or exec_p <= 0: continue
        pos = portfolio.pop(code)
        proceeds = pos['shares'] * exec_p
        pnl_amt = pos['shares'] * (exec_p - pos['entry_price'])
        pnl_pct_actual = (exec_p - pos['entry_price']) / pos['entry_price'] * 100
        cash += proceeds
        trades.append({
            'ID': trade_id, '日期': date, '代码': code,
            '买卖方向': '卖出',
            '价格': round(exec_p, 2), '数量': pos['shares'],
            '金额': round(proceeds, 0),
            '浮盈亏%': round(pnl_pct_actual, 2),
            '持仓天数': hd, '卖出原因': reason,
            '买入价格': round(pos['entry_price'], 2),
        })
        trade_id += 1

    # 3. BUY: 计算 total_nav (NAV用开盘价估值, 开盘时不知道收盘价)
    port_value = 0
    for code, pos in portfolio.items():
        open_p = get_open(code, date)
        if open_p is not None and open_p > 0:
            port_value += pos['shares'] * open_p
        else:
            port_value += pos['shares'] * pos['entry_price']
    total_nav = cash + port_value

    # 4. 买入: 用开盘价排序, 以开盘价执行 (涨停过滤)
    while len(portfolio) < max_pos:
        bought = False
        for score, code, price_open, bar_pct_v, dif_pct_v, slope_v, br_v in top_pool_sorted:
            if code in portfolio: continue
            if cash < 50000: break
            # 涨停检查: 开盘涨停 → 买不进去, 跳过
            ind_b = IND.get(code)
            if ind_b:
                i_b = ind_b['date_to_idx'].get(date)
                if i_b is not None and i_b > 0:
                    prev_close_b = ind_b['prices'][i_b-1]
                    if prev_close_b and prev_close_b > 0 and is_limit_up_open(code, price_open, prev_close_b):
                        continue  # 开盘涨停, 买不进
            exec_p = price_open  # 用开盘价执行
            if exec_p is None or exec_p <= 0: continue
            shares = int(POS_SIZE * total_nav / exec_p / 100) * 100
            if shares < 100: continue
            cost = shares * exec_p
            if cost > cash:
                shares = int(cash / exec_p / 100) * 100
                cost = shares * exec_p
            if shares < 100: continue
            cash -= cost
            portfolio[code] = {
                'shares': shares, 'entry_price': exec_p,
                'entry_date': date, 'entry_di': di, 'score': score,
            }
            trades.append({
                'ID': trade_id, '日期': date, '代码': code, '买卖方向': '买入',
                '价格': round(exec_p, 2), '数量': shares, '金额': round(cost, 0),
                '浮盈亏%': 0, '持仓天数': 0, '卖出原因': '', '买入价格': round(exec_p, 2),
            })
            trade_id += 1
            bought = True
            break
        if not bought: break

    # 5. 记录每日净值 (用收盘价估值)
    port_value = 0
    for code, pos in portfolio.items():
        ind = IND[code]
        i = ind['date_to_idx'].get(date)
        if i is not None:
            p = ind['prices'][i]
            port_value += pos['shares'] * (p if (p is not None and p > 0) else pos['entry_price'])
    daily_nav.append({
        '日期': date, '净值': round((cash + port_value) / INIT_CASH, 6),
        '持仓数量': len(portfolio), '现金': round(cash, 0),
        '持仓市值': round(port_value, 0)
    })

    if di % 200 == 0:
        print(f"  [{di+1}/{n}] {date} 持仓{len(portfolio)}只 净值{(cash+port_value)/INIT_CASH:.4f} 候选{len(candidates)}只")
        sys.stdout.flush()

# ===== 最终结算 =====
print("\n最终结算...")
sys.stdout.flush()
last_date = sim_dates[-1]
last_di = len(sim_dates) - 1
for code, pos in list(portfolio.items()):
    ind = IND[code]
    i = ind['date_to_idx'].get(last_date)
    if i is None: continue
    close_p = ind['prices'][i]
    if close_p is None or close_p <= 0: continue
    exec_p = get_open(code, last_date)
    if exec_p is None or exec_p <= 0: continue
    pnl_pct = (exec_p - pos['entry_price']) / pos['entry_price']
    pnl_amt = pos['shares'] * (exec_p - pos['entry_price'])
    cash += pos['shares'] * exec_p
    hold_days = last_di - pos['entry_di']
    trades.append({
        'ID': trade_id, '日期': last_date, '代码': code,
        '买卖方向': '卖出',
        '价格': round(exec_p, 2), '数量': pos['shares'],
        '金额': round(pos['shares'] * exec_p, 0),
        '浮盈亏%': round(pnl_pct*100, 2),
        '持仓天数': hold_days, '卖出原因': '最终结算',
        '买入价格': round(pos['entry_price'], 2),
    })
    trade_id += 1
    portfolio.pop(code)

# ===== 统计 =====
final_nav = cash / INIT_CASH
print(f"\n{'='*60}")
print(f"模拟盘 纯死叉版 (回测引擎, {START_DATE}~{END_DATE})")
print(f"最终净值: {final_nav:.4f}  ({(final_nav-1)*100:+.2f}%)")
print(f"期末现金: {cash:,.0f}")

wins = [t for t in trades if t.get('浮盈亏%', 0) > 0]
losts = [t for t in trades if t.get('浮盈亏%', 0) <= 0]
print(f"总交易: {len(trades)}, 胜率: {len(wins)/max(1,len(trades))*100:.1f}%")
print(f"盈利笔均: {sum(t.get('浮盈亏%',0) for t in wins)/max(1,len(wins)):+,.2f}%")
print(f"亏损笔均: {sum(t.get('浮盈亏%',0) for t in losts)/max(1,len(losts)):+,.2f}%")

# 最大回撤 (running peak)
navs = [d['净值'] for d in daily_nav]
peak = navs[0]; max_dd = 0
for nv in navs:
    if nv > peak: peak = nv
    dd = (peak - nv) / peak
    if dd > max_dd: max_dd = dd
print(f"最大回撤: {max_dd*100:.2f}%")

# 年度收益
year_returns = {}
year_start = {}
for d in daily_nav:
    y = d['日期'][:4]
    if y not in year_start:
        year_start[y] = d['净值']
    year_returns[y] = d['净值']
print(f"\n年度收益:")
prev = 1.0
for y in sorted(year_returns.keys()):
    yr = (year_returns[y] / prev - 1) * 100
    print(f"  {y}: {yr:+.2f}%")
    prev = year_returns[y]

# 卖出原因分布
reasons = Counter(t['卖出原因'] for t in trades)
print(f"\n卖出原因:")
for r, c in sorted(reasons.items(), key=lambda x: -x[1]):
    print(f"  {r}: {c}笔 ({c/len(trades)*100:.1f}%)")

# 平均持仓天数
print(f"平均持仓天数: {sum(t['持仓天数'] for t in trades)/max(1,len(trades)):.1f}")

# 保存结果
result = {
    'final_nav': final_nav,
    'final_cash': cash,
    'return_pct': (final_nav-1)*100,
    'max_dd_pct': max_dd*100,
    'n_trades': len(trades),
    'win_rate': len(wins)/max(1,len(trades))*100,
    'year_returns': {y: (year_returns[y]/year_start[y]-1)*100 for y in year_returns},
    'reasons': dict(reasons),
    'avg_hold_days': sum(t['持仓天数'] for t in trades)/max(1,len(trades)),
    'daily_nav': daily_nav,
    'trades': trades,
}
out_file = f'/workspace/模拟盘_纯死叉版_固定_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
with open(out_file, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print(f"\n结果已保存: {out_file}")

# ===== 输出Excel (兼容原模拟盘格式) =====
print("写出Excel...", flush=True)
import openpyxl as _xl_out
from openpyxl.styles import Font as _Font, Alignment as _Align

wb = _xl_out.Workbook()
# Sheet1: 当前持仓
ws1 = wb.active; ws1.title = '当前持仓'
h1 = ['代码', '持仓数量', '买入价格', '盈亏%', '市值']
for ci, h in enumerate(h1, 1):
    ws1.cell(row=1, column=ci, value=h)
port_list = sorted(portfolio.items(), key=lambda x: x[1].get('entry_price', 0), reverse=True)
for ri, (code, pos) in enumerate(port_list, 2):
    pnl = (daily_nav[-1]['净值'] * INIT_CASH / pos['entry_price'] - 1) * 100 if pos['entry_price'] > 0 else 0
    ws1.cell(row=ri, column=1, value=code)
    ws1.cell(row=ri, column=2, value=pos['shares'])
    ws1.cell(row=ri, column=3, value=round(pos['entry_price'], 2))
    ws1.cell(row=ri, column=4, value=round(pnl, 2))
    ws1.cell(row=ri, column=5, value=round(daily_nav[-1]['净值'] * INIT_CASH / max(len(port_list),1), 0))
ws1.freeze_panes = 'A2'

# Sheet2: 交易记录
ws2 = wb.create_sheet('交易记录')
h2 = ['ID', '日期', '代码', '买卖方向', '价格', '数量', '金额', '盈亏%', '持仓天数', '卖出原因', '买入价格']
for ci, h in enumerate(h2, 1):
    ws2.cell(row=1, column=ci, value=h)
for ri, t in enumerate(trades, 2):
    ws2.cell(row=ri, column=1, value=t.get('ID', ri-1))
    ws2.cell(row=ri, column=2, value=t.get('日期', ''))
    ws2.cell(row=ri, column=3, value=t.get('代码', ''))
    ws2.cell(row=ri, column=4, value='买入' if t.get('持仓天数', 0) == 0 else '卖出')
    ws2.cell(row=ri, column=5, value=t.get('价格', 0))
    ws2.cell(row=ri, column=6, value=t.get('数量', 0))
    ws2.cell(row=ri, column=7, value=t.get('金额', 0))
    ws2.cell(row=ri, column=8, value=round(t.get('浮盈亏%', 0), 2))
    ws2.cell(row=ri, column=9, value=t.get('持仓天数', 0))
    ws2.cell(row=ri, column=10, value=t.get('卖出原因', ''))
    ws2.cell(row=ri, column=11, value=t.get('买入价格', 0))
ws2.freeze_panes = 'A2'

# Sheet3: 净值曲线
ws3 = wb.create_sheet('净值曲线')
h3 = ['日期', '净值', '持仓']
for ci, h in enumerate(h3, 1):
    ws3.cell(row=1, column=ci, value=h)
for ri, d in enumerate(daily_nav, 2):
    ws3.cell(row=ri, column=1, value=d['日期'])
    ws3.cell(row=ri, column=2, value=d['净值'])
    ws3.cell(row=ri, column=3, value=d.get('持仓数量', ''))
ws3.freeze_panes = 'A2'

# Sheet4: 统计分析
ws4 = wb.create_sheet('统计分析')
stats = [
    {'指标': '总收益%', '数值': round((final_nav-1)*100, 2)},
    {'指标': '最大回撤%', '数值': round(max_dd*100, 2)},
    {'指标': '胜率%', '数值': round(len([t for t in trades if t.get('浮盈亏%',0)>0])/max(1,len(trades))*100, 1)},
    {'指标': '总交易', '数值': len(trades)},
    {'指标': '平均持仓天数', '数值': round(sum(t['持仓天数'] for t in trades)/max(1,len(trades)), 1)},
    {'指标': '最终净值', '数值': round(final_nav, 4)},
]
for r, c in sorted(reasons.items(), key=lambda x: -x[1]):
    stats.append({'指标': f'卖出:{r}', '数值': c})
for ci, h in enumerate(['指标', '数值'], 1):
    ws4.cell(row=1, column=ci, value=h)
for ri, s in enumerate(stats, 2):
    ws4.cell(row=ri, column=1, value=s['指标'])
    ws4.cell(row=ri, column=2, value=s['数值'])
ws4.freeze_panes = 'A2'

xlsx_file = out_file.replace('.json', '.xlsx')
wb.save(xlsx_file)
print(f"Excel: {xlsx_file}")
print(f"  Sheet1 当前持仓: {len(port_list)}行")
print(f"  Sheet2 交易记录: {len(trades)}行")
print(f"  Sheet3 净值曲线: {len(daily_nav)}行")
print(f"  Sheet4 统计分析: {len(stats)}行")
