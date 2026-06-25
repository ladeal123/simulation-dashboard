"""模拟盘 V46_v62_v20_final: T日收盘信号 + T+1开盘执行 + 量比加分
策略: MACD(10,20,9) 3日金叉选股
- 选股: BAR>0, 价格>MA144, 3日内金叉
- 排序: (DIF%x0.3 + BAR%x0.7) - 0.2xvol10
- 量比加分: 5日均量比 >=1.0/1.2/1.5 -> +10%/+20%/+35%
- 金叉衰减: gc_days=0(1.0), >=1(0.5)
- 卖出: 止损-15%, 死叉, 动态止盈(浮盈>50%且<MA5)
- 仓位: bull=50, flat=20, short=5
- 数据: 后复权 + 成交量
- 交易记录: 按日期降序(最新在前)
"""
import sys, math, datetime
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===================== 参数 =====================
FAST, SLOW, SIGNAL = 10, 20, 9           # 个股 MACD (v62)
INDEX_FAST, INDEX_SLOW, INDEX_DEA = 10, 20, 9  # 中证1000 V4 MACD
INIT_CASH = 10_000_000
POS_SIZE = 0.02
MIN_HOLD = 3
STOP_LOSS = -0.15                         # 止损 -15%
PROFIT_THRESHOLD = 0.50                   # 动态止盈 50%
VOL_ALPHA = 0.20
# 量比加分阈值
VOL_BONUS_MAP = [(1.5, 0.35), (1.2, 0.20), (1.0, 0.10)]
# 金叉衰减
GC_PENALTY = {0: 1.0, 1: 0.5, 2: 0.5, 3: 0.5}
# V4 状态机 (无平滑)
STATE_TARGET = {'bull': 50, 'flat': 20, 'short': 5}
TH_SHORT = 0.0
BOTTOM = -3.0
# 交易成本
COMMISSION_RATE = 0.0001
STAMP_TAX_RATE = 0.0005
# 数据起止
BACKTEST_START = '2026-01-01'
# 文件路径
POOL_FILE = '/workspace/股票池.xlsx'
CLOSE_FILE = '/workspace/模拟盘收盘价后复权.xlsx'
OPEN_FILE = '/workspace/模拟盘开盘价后复权.xlsx'
INDEX_FILE = '/workspace/模拟盘指数价格数据.xlsx'
VOL_FILE = '/workspace/模拟盘成交量.xlsx'


# ===================== 工具函数 =====================
def ema_v46(values, n_):
    a = 2.0 / (n_ + 1)
    out = [None] * len(values)
    fv = None
    for i, v in enumerate(values):
        if v is not None and v > 0:
            fv = i
            break
    if fv is None:
        return out
    out[fv] = values[fv]
    for i in range(fv + 1, len(values)):
        v = values[i]
        if v is None or v == 0:
            out[i] = out[i - 1]
        else:
            out[i] = a * v + (1 - a) * out[i - 1]
    return out


def dif_slope(dif, win=2):
    out = [0.0] * len(dif)
    for i in range(win, len(dif)):
        y = [dif[i - j] for j in range(win - 1, -1, -1)]
        xm = [j - (win - 1) / 2 for j in range(win)]
        ym = [yi - sum(y) / win for yi in y]
        num = sum(xm[j] * ym[j] for j in range(win))
        den = sum(xm[j] ** 2 for j in range(win))
        out[i] = num / den if den != 0 else 0
    return out


def calc_v4_state_v46(idx_dates, prices, dif, dif_pct, dea, slope, bar,
                      th_short=0.0, bottom=-3.0):
    pos = 0
    result = []
    prev_bar = None
    for i in range(1, len(idx_dates)):
        rd = dif[i] if dif[i] is not None else 0
        rdea = dea[i] if dea[i] is not None else 0
        pd_ = dif[i - 1] if dif[i - 1] is not None else 0
        pdea = dea[i - 1] if dea[i - 1] is not None else 0
        rdp = dif_pct[i]
        rsl = slope[i]
        rbar = bar[i] if (bar is not None and i < len(bar) and bar[i] is not None) else 0
        gcu = pd_ <= pdea and rd > rdea
        gcd = pd_ >= pdea and rd < pdea
        if pos == 0:
            if gcu:
                pos = 1
            elif rdp > 0 and rsl > 0.02:
                pos = 1
            elif rdp < bottom and rsl > 0.03:
                pos = 1
            elif rdp <= th_short:
                pos = -1
        elif pos == 1:
            if gcd:
                pos = 0
                if rdp <= th_short:
                    pos = -1
            elif prev_bar is not None and prev_bar < 0:
                if rdp > 0 and rsl < -0.02:
                    pos = 0
        elif pos == -1:
            if gcu:
                pos = 1
            elif rdp < bottom and rsl > 0.03:
                pos = 1
        result.append(pos)
        prev_bar = rbar
    return result


def calc_ema_s(pl, p_):
    """EMA 计算 (支持nan)"""
    n_ = len(pl)
    a = 2.0 / (p_ + 1)
    out = [float('nan')] * n_
    fv_ = next((i for i, v in enumerate(pl) if v > 0), None)
    if fv_ is None:
        return out
    out[fv_] = pl[fv_]
    for i in range(fv_ + 1, n_):
        if pl[i] == 0:
            out[i] = out[i - 1] if not math.isnan(out[i - 1]) else float('nan')
        else:
            out[i] = a * pl[i] + (1 - a) * out[i - 1]
    return out


# ===================== 加载股票池 =====================
print("[1/7] 加载股票池...", flush=True)
code_to_name = {}
code_to_industry = {}
wb_pool = openpyxl.load_workbook(POOL_FILE, data_only=True, read_only=True)
ws_pool = wb_pool['股票池']
for row in ws_pool.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    code = str(row[0]).strip()
    name = str(row[1]).strip() if row[1] else ''
    ind = str(row[2]).strip() if row[2] else ''
    if code and ind:
        code_to_name[code] = name
        code_to_industry[code] = ind
wb_pool.close()
print(f"  股票池: {len(code_to_industry)} 只")


# ===================== 加载收盘价 + 开盘价 =====================
print("[2/7] 加载收盘价/开盘价(后复权)...", flush=True)


def load_price_file(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['close_price']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    codes = [str(c).strip() for c in rows[0][1:] if c is not None]
    names = [str(n).strip() if n else '' for n in rows[1][1:1 + len(codes)]]
    price_dict = defaultdict(dict)
    for row in rows[4:]:
        if not row or row[0] is None:
            continue
        d = row[0]
        if isinstance(d, datetime.datetime):
            ds = d.strftime('%Y-%m-%d')
        else:
            ds = str(d)[:10]
        if not ds.startswith('20'):
            continue
        for i, code in enumerate(codes):
            if i + 1 >= len(row):
                break
            v = row[i + 1]
            if v is not None and isinstance(v, (int, float)) and v > 0:
                price_dict[code][ds] = v
    return codes, names, sorted(price_dict.keys()), price_dict


close_codes, close_names, _, close_data = load_price_file(CLOSE_FILE)
open_codes, open_names, _, open_data = load_price_file(OPEN_FILE)
print(f"  收盘价: {len(close_codes)} 只")
print(f"  开盘价: {len(open_codes)} 只")

# 名称补全
for c, n in zip(close_codes, close_names):
    if c not in code_to_name and n:
        code_to_name[c] = n

all_dates = sorted({d for c in close_data for d in close_data[c]})
print(f"  日期范围: {all_dates[0]} ~ {all_dates[-1]}, 共 {len(all_dates)} 天", flush=True)


# ===================== 加载成交量 =====================
print("[3/7] 加载成交量...", flush=True)


def load_vol_file(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['close_price']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    codes = [str(c).strip() for c in rows[0][1:] if c is not None]
    vol_dict = defaultdict(dict)
    for row in rows[4:]:
        if not row or row[0] is None:
            continue
        d = row[0]
        if isinstance(d, datetime.datetime):
            ds = d.strftime('%Y-%m-%d')
        else:
            ds = str(d)[:10]
        if not ds.startswith('20'):
            continue
        for i, code in enumerate(codes):
            if i + 1 >= len(row):
                break
            v = row[i + 1]
            if v is not None and isinstance(v, (int, float)) and v > 0:
                vol_dict[code][ds] = v
    return codes, vol_dict


vol_codes, vol_data = load_vol_file(VOL_FILE)
common_codes = set(vol_codes) & set(close_codes)
print(f"  成交量: {len(vol_codes)} 只, 与收盘价共有: {len(common_codes)} 只")

# 从股票池Excel加载指数成分股归属和模拟盘池
print("[3.5/7] 加载股票池成分信息...", flush=True)
pool_wb = openpyxl.load_workbook(POOL_FILE, data_only=True, read_only=True)

# 沪深300 / 中证500 / 中证1000 成分股
index_sets = {}
for sheet_name, tag in [('沪深300', '300'), ('中证500', '500'), ('中证1000', '1000')]:
    ws = pool_wb[sheet_name]
    codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            codes.add(str(row[0]).strip())
    index_sets[tag] = codes
    print(f"  {sheet_name}: {len(codes)}只")

# 模拟盘池 (用于判断池内/池外)
pool_ws = pool_wb['模拟盘']
pool_codes = set()
for row in pool_ws.iter_rows(min_row=2, values_only=True):
    if row and row[0]:
        pool_codes.add(str(row[0]).strip())
print(f"  模拟盘池: {len(pool_codes)}只")
pool_wb.close()

def get_index_tag(code):
    raw = code.strip()
    for tag in ['300', '500', '1000']:
        if raw in index_sets[tag]:
            return tag
    return '其他'

def is_in_pool(code):
    return '是' if code.strip() in pool_codes else '否'


# ===================== 加载指数 + V4状态 =====================
print("[4/7] 加载中证1000指数并计算V4状态...", flush=True)


def load_index(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['index']
    rows = list(ws.iter_rows(min_row=6, values_only=True))
    wb.close()
    idx_data = {}
    for row in rows:
        if not row or row[0] is None:
            continue
        d = row[0]
        if isinstance(d, datetime.datetime):
            ds = d.strftime('%Y-%m-%d')
        else:
            ds = str(d).split()[0]
        if not ds.startswith('20'):
            continue
        if len(row) > 1 and row[1] is not None and isinstance(row[1], (int, float)) and row[1] > 0:
            idx_data[ds] = float(row[1])
    return idx_data


idx_data = load_index(INDEX_FILE)
print(f"  指数: {len(idx_data)} 天", flush=True)

idx_dates = sorted(idx_data.keys())
idx_prices = [idx_data[d] for d in idx_dates]
idx_ef = ema_v46(idx_prices, INDEX_FAST)
idx_es = ema_v46(idx_prices, INDEX_SLOW)
idx_dif = [idx_ef[i] - idx_es[i] if (idx_ef[i] is not None and idx_es[i] is not None) else None
           for i in range(len(idx_prices))]
idx_dif_pct = [idx_dif[i] / idx_prices[i] * 100 if (idx_dif[i] is not None and idx_prices[i] > 0) else 0.0
               for i in range(len(idx_prices))]
idx_dea = ema_v46(idx_dif, INDEX_DEA)
idx_bar = [(idx_dif[i] - idx_dea[i]) * 2 if (idx_dif[i] is not None and idx_dea[i] is not None) else None
           for i in range(len(idx_prices))]
idx_slope = dif_slope(idx_dif_pct, 2)
v4_states = calc_v4_state_v46(idx_dates, idx_prices, idx_dif, idx_dif_pct, idx_dea,
                              idx_slope, idx_bar, TH_SHORT, BOTTOM)
idx_state_map = {idx_dates[i + 1]: v4_states[i] for i in range(len(v4_states))}
print(f"  V4 状态: {len(idx_state_map)} 天", flush=True)


# ===================== 预计算个股指标 =====================
print("[5/7] 预计算个股 MACD/MA144/vol10...", flush=True)
IND = {}
for code, dp in close_data.items():
    d_list = sorted(dp.keys())
    if len(d_list) < 150:
        continue
    pl = [dp.get(d, 0) for d in d_list]
    # MACD(10,20,9)
    ef = calc_ema_s(pl, FAST)
    es = calc_ema_s(pl, SLOW)
    dif = [ef[i] - es[i] if not (math.isnan(ef[i]) or math.isnan(es[i])) else float('nan')
           for i in range(len(pl))]
    dea = calc_ema_s(dif, SIGNAL)
    bar = [(dif[i] - dea[i]) * 2.0 if not (math.isnan(dif[i]) or math.isnan(dea[i])) else float('nan')
           for i in range(len(dif))]
    # MA5, MA144
    ma5 = calc_ema_s(pl, 5)
    ma144 = calc_ema_s(pl, 144)
    # vol10
    vol10 = [float('nan')] * len(pl)
    for j in range(10, len(pl)):
        window = [(pl[k] / pl[k - 1] - 1) * 100 if pl[k - 1] > 0 and pl[k] > 0 else 0
                  for k in range(j - 9, j + 1)]
        if len(window) == 10:
            mean_w = sum(window) / 10
            var_w = sum((x - mean_w) ** 2 for x in window) / 10
            vol10[j] = var_w ** 0.5
    date_to_idx = {d: i for i, d in enumerate(d_list)}
    IND[code] = {
        'dates': d_list, 'date_to_idx': date_to_idx,
        'prices': pl,
        'dif': dif, 'dea': dea, 'bar': bar,
        'ma5': ma5, 'ma144': ma144, 'vol10': vol10,
    }
print(f"  有效股票: {len(IND)} 只", flush=True)


def get_close(code, date):
    dp = close_data.get(code)
    if dp is None:
        return None
    return dp.get(date)


def get_open(code, date):
    dp = open_data.get(code)
    if dp is None:
        return None
    return dp.get(date)


def get_board(code):
    if code.startswith('30') or code.startswith('688') or code.startswith('301'):
        return 'chinext'
    return 'main'


def calc_vol_bonus(code, signal_date, all_dates_sorted):
    """量比加分: 信号日成交量/5日均量, 返回 (bonus, ratio)"""
    vd = vol_data.get(code)
    if vd is None:
        return 0.0, 0.0
    v_signal = vd.get(signal_date)
    if v_signal is None or v_signal <= 0:
        return 0.0, 0.0
    try:
        idx = all_dates_sorted.index(signal_date)
    except ValueError:
        return 0.0, 0.0
    sum5 = 0.0; cnt5 = 0
    for off in range(1, 6):
        if idx - off < 0:
            continue
        v = vd.get(all_dates_sorted[idx - off])
        if v is not None and v > 0:
            sum5 += v; cnt5 += 1
    if cnt5 < 3 or sum5 <= 0:
        return 0.0, 0.0
    ratio = v_signal / (sum5 / cnt5)
    for threshold, bonus in VOL_BONUS_MAP:
        if ratio >= threshold:
            return bonus, round(ratio, 2)
    return 0.0, round(ratio, 2)


# ===================== 主回测 =====================
print("[6/7] 开始回测 (T日收盘信号 + T+1开盘执行 + 量比加分)...", flush=True)
sim_dates = [d for d in all_dates if d >= BACKTEST_START]
print(f"  回测区间: {sim_dates[0]} ~ {sim_dates[-1]}, 共 {len(sim_dates)} 天", flush=True)

portfolio = {}
cash = INIT_CASH
trades = []
daily_nav = []
daily_top50 = []
last_full_pool = []
daily_rank_up = []  # 每日排名上升最快Top10
prev_day_rank = {}  # 前一日排名(code->rank)
trade_id = 0
n = len(sim_dates)
total_commission = 0.0
total_stamp_tax = 0.0
ytb_skip_count = 0
sell_reason_counter = Counter()
vol_bonus_count = {0.10: 0, 0.20: 0, 0.35: 0}

for di, date in enumerate(sim_dates):
    exec_date = sim_dates[di + 1] if di + 1 < n else date
    signal_date = date  # T 日

    # V4 状态
    idx_state_raw = idx_state_map.get(date, 1)
    state_key = 'bull' if idx_state_raw == 1 else ('short' if idx_state_raw == -1 else 'flat')
    max_pos = STATE_TARGET.get(state_key, STATE_TARGET['bull'])

    # 1. 候选池 (T日收盘信号)
    candidates = []
    for code, ind in IND.items():
        if code in portfolio:
            continue
        i = ind['date_to_idx'].get(signal_date)
        if i is None:
            continue
        c_i = ind['prices'][i]
        if c_i <= 0 or math.isnan(c_i):
            continue
        dif_i = ind['dif'][i]
        dea_i = ind['dea'][i]
        bar_i = ind['bar'][i]
        ma144_i = ind['ma144'][i]
        if math.isnan(dif_i) or math.isnan(dea_i) or math.isnan(ma144_i):
            continue
        if math.isnan(bar_i) or bar_i <= 0:
            continue
        if c_i < ma144_i:
            continue
        # 3日内金叉
        gc_days = -1
        for off in range(3):
            if i - off < 0:
                continue
            dn = ind['dif'][i - off]
            dp = ind['dif'][i - off - 1] if i - off - 1 >= 0 else float('nan')
            en = ind['dea'][i - off]
            ep = ind['dea'][i - off - 1] if i - off - 1 >= 0 else float('nan')
            if not (math.isnan(dn) or math.isnan(dp) or math.isnan(en) or math.isnan(ep)):
                if dp <= ep and dn > en:
                    gc_days = off
                    break
        if gc_days < 0:
            continue
        # T+1开盘价
        j = ind['date_to_idx'].get(exec_date) if exec_date != signal_date else i
        if j is None:
            continue
        o_j = get_open(code, exec_date)
        if o_j is None or o_j <= 0:
            o_j = c_i
        # 涨停过滤
        limit_p = round(c_i * (1.095 if get_board(code) == 'main' else 1.195), 2)
        is_zt = o_j >= limit_p
        # 打分: (DIF% x0.3 + BAR% x0.7) - 0.2 x vol10
        dif_pct_v = dif_i / c_i * 100
        bar_pct_v = bar_i / c_i * 100
        vol10_v = ind['vol10'][i] if i < len(ind['vol10']) and not math.isnan(ind['vol10'][i]) else 0
        base_score = dif_pct_v * 0.3 + bar_pct_v * 0.7 - VOL_ALPHA * vol10_v
        # 量比加分
        vol_bonus, vol_ratio = calc_vol_bonus(code, signal_date, all_dates)
        score = base_score * GC_PENALTY.get(gc_days, 0.2) * (1 + vol_bonus)
        if vol_bonus > 0:
            vol_bonus_count[vol_bonus] = vol_bonus_count.get(vol_bonus, 0) + 1
        candidates.append({
            'code': code, 'score': score, 'gc_days': gc_days,
            'o_j': o_j, 'is_zt': is_zt, 'c_i': c_i,
            'dif_pct': dif_pct_v, 'bar_pct': bar_pct_v, 'vol10': vol10_v,
            'vol_bonus': vol_bonus, 'vol_ratio': vol_ratio,
        })

    # 排序
    candidates.sort(key=lambda x: -x['score'])

    # 记录每日Top50
    for rank, c in enumerate(candidates[:50], 1):
        daily_top50.append({
            '日期': signal_date, '执行日': exec_date, '排名': rank, '代码': c['code'],
            '名称': code_to_name.get(c['code'], ''),
            '指数归属': get_index_tag(c['code']),
            '池内': is_in_pool(c['code']),
            '得分': round(c['score'], 4),
            '金叉天数': c['gc_days'],
            'DIF%': round(c['dif_pct'], 4),
            'BAR%': round(c['bar_pct'], 4),
            'vol10': round(c['vol10'], 4),
            '量比': c['vol_ratio'],
            '量比加分': c['vol_bonus'],
            '涨停': '是' if c['is_zt'] else '',
        })
    last_full_pool = [{
        '日期': signal_date, '执行日': exec_date, '排名': idx + 1, '代码': c['code'],
        '名称': code_to_name.get(c['code'], ''),
        '指数归属': get_index_tag(c['code']),
        '池内': is_in_pool(c['code']),
        '得分': round(c['score'], 4),
        '金叉天数': c['gc_days'],
        'DIF%': round(c['dif_pct'], 4),
        'BAR%': round(c['bar_pct'], 4),
        'vol10': round(c['vol10'], 4),
        '量比': c['vol_ratio'],
        '量比加分': c['vol_bonus'],
        '涨停': '是' if c['is_zt'] else '',
    } for idx, c in enumerate(candidates)]
    
    # 记录每日排名上升最快的10只
    if prev_day_rank:
        rank_changes = []
        for idx, c in enumerate(candidates[:100]):  # 只看前100名
            code = c['code']
            cur_rank = idx + 1
            prev_rank = prev_day_rank.get(code)
            if prev_rank is not None:
                change = prev_rank - cur_rank  # 正数=排名上升
                if change > 0:
                    rank_changes.append({
                        '日期': signal_date,
                        '代码': code,
                        '名称': code_to_name.get(code, ''),
                        '指数归属': get_index_tag(code),
                        '当日排名': cur_rank,
                        '昨日排名': prev_rank,
                        '排名变化': change,
                        '得分': round(c['score'], 4),
                        '量比': c['vol_ratio'],
                    })
        # 取上升最快的Top10
        rank_changes.sort(key=lambda x: -x['排名变化'])
        daily_rank_up.extend(rank_changes[:10])
    # 更新前一日排名（只保留Top200）
    prev_day_rank = {c['code']: idx + 1 for idx, c in enumerate(candidates[:200])}
    
    # 2. 卖出
    to_sell = []
    for code, pos in list(portfolio.items()):
        ind = IND[code]
        i = ind['date_to_idx'].get(signal_date)
        if i is None:
            continue
        o_j = get_open(code, exec_date)
        if o_j is None or o_j <= 0:
            o_j = ind['prices'][i]
            if o_j <= 0:
                continue
        hold_days = di - pos['entry_di']
        pnl_pct = (o_j - pos['entry_price']) / pos['entry_price']
        # max_price
        c_i = ind['prices'][i]
        if c_i is not None and c_i > 0 and not math.isnan(c_i):
            if c_i > pos.get('max_price', 0):
                pos['max_price'] = c_i
        # 止损
        if hold_days >= MIN_HOLD and pnl_pct <= STOP_LOSS:
            to_sell.append((code, o_j, pnl_pct, '止损'))
            continue
        # 死叉
        if hold_days >= MIN_HOLD and i >= 1:
            dif_i = ind['dif'][i]
            dea_i = ind['dea'][i]
            dif_prev = ind['dif'][i - 1]
            dea_prev = ind['dea'][i - 1]
            if not (math.isnan(dif_i) or math.isnan(dea_i) or math.isnan(dif_prev) or math.isnan(dea_prev)):
                if dif_prev > dea_prev and dif_i <= dea_i:
                    to_sell.append((code, o_j, pnl_pct, '死叉'))
                    continue
        # 动态止盈
        if pnl_pct > PROFIT_THRESHOLD:
            ma5_i = ind['ma5'][i]
            if not math.isnan(ma5_i) and c_i < ma5_i:
                to_sell.append((code, o_j, pnl_pct, '动态止盈'))
                continue
    # 超仓位强卖
    if len(portfolio) > max_pos:
        port_list = [{'code': c, 'score': p.get('score', 0), 'entry_di': p['entry_di']}
                     for c, p in portfolio.items()]
        port_list.sort(key=lambda x: x['score'])
        for item in port_list[:len(portfolio) - max_pos]:
            code = item['code']
            o_j = get_open(code, exec_date)
            if o_j is None or o_j <= 0:
                continue
            pnl = (o_j - portfolio[code]['entry_price']) / portfolio[code]['entry_price']
            to_sell.append((code, o_j, pnl, f'超仓位({max_pos})'))
    # 执行卖出 (去重)
    sold_set = set()
    for code, price, pnl_pct_signal, reason in to_sell:
        if code not in portfolio or code in sold_set:
            continue
        sold_set.add(code)
        exec_p = get_open(code, exec_date)
        if exec_p is None or exec_p <= 0:
            continue
        pos = portfolio.pop(code)
        cost_tax = exec_p * pos['shares'] * (COMMISSION_RATE + STAMP_TAX_RATE)
        total_commission += exec_p * pos['shares'] * COMMISSION_RATE
        total_stamp_tax += exec_p * pos['shares'] * STAMP_TAX_RATE
        proceeds = exec_p * pos['shares'] - cost_tax
        pnl_amt = pos['shares'] * (exec_p - pos['entry_price']) - cost_tax
        pnl_pct_actual = (exec_p - pos['entry_price']) / pos['entry_price'] * 100
        cash += proceeds
        trades.append({
            'ID': trade_id, '日期': exec_date, '代码': code,
            '名称': code_to_name.get(code, ''),
            '买卖方向': '卖出',
            '价格': round(exec_p, 4),
            '数量': pos['shares'],
            '金额': round(proceeds, 2),
            '盈亏%': round(pnl_pct_actual, 2),
            '盈亏金额': round(pnl_amt, 2),
            '持仓天数': di - pos['entry_di'],
            '卖出原因': reason,
            '信号日期': signal_date,
            '买入日期': pos['entry_date'],
            '买入价格': round(pos['entry_price'], 4),
        })
        trade_id += 1
        sell_reason_counter[reason] += 1

    # 3. 买入
    port_value = 0
    for code, pos in portfolio.items():
        open_p = get_open(code, exec_date)
        if open_p is not None and open_p > 0:
            port_value += pos['shares'] * open_p
        else:
            port_value += pos['shares'] * pos['entry_price']
    total_nav = cash + port_value

    for c in candidates:
        if len(portfolio) >= max_pos or cash <= 50000:
            break
        code = c['code']
        if code in portfolio:
            continue
        if c['is_zt']:
            ytb_skip_count += 1
            continue
        exec_p = get_open(code, exec_date)
        if exec_p is None or exec_p <= 0:
            continue
        shares = max(int(POS_SIZE * total_nav / exec_p / 100) * 100, 100)
        cost = shares * exec_p
        total_cost = cost * (1 + COMMISSION_RATE)
        if total_cost > cash * 0.95:
            continue
        total_commission += cost * COMMISSION_RATE
        cash -= total_cost
        portfolio[code] = {
            'shares': shares, 'entry_price': exec_p,
            'entry_date': exec_date, 'entry_di': di,
            'max_price': exec_p, 'score': c['score'],
        }
        trades.append({
            'ID': trade_id, '日期': exec_date, '代码': code,
            '名称': code_to_name.get(code, ''),
            '买卖方向': '买入',
            '价格': round(exec_p, 4),
            '数量': shares,
            '金额': round(cost, 2),
            '盈亏%': 0.0,
            '盈亏金额': 0.0,
            '持仓天数': 0, '卖出原因': '',
            '信号日期': signal_date,
            '买入日期': exec_date,
            '买入价格': round(exec_p, 4),
        })
        trade_id += 1

    # 4. NAV估值
    port_value = 0
    for code, pos in portfolio.items():
        close_p = get_close(code, date)
        if close_p is not None and close_p > 0:
            port_value += pos['shares'] * close_p
        else:
            port_value += pos['shares'] * pos['entry_price']
    nav = (cash + port_value) / INIT_CASH
    daily_nav.append({
        '日期': date, '净值': round(nav, 6),
        '持仓数量': len(portfolio), '现金': round(cash, 2),
        '持仓市值': round(port_value, 2),
    })

    if di % 20 == 0 or di == n - 1:
        print(f"  [{di + 1}/{n}] {date}(执行{exec_date}) {state_key}持仓{len(portfolio)}只 净值{nav:.4f}", flush=True)


# ===================== 最终结算 =====================
print("[7/7] 最终结算 + 输出Excel (交易记录按日期降序)...", flush=True)
last_date = sim_dates[-1]
last_di = len(sim_dates) - 1
final_holdings = []
for code, pos in portfolio.items():
    close_p = get_close(code, last_date)
    if close_p is None or close_p <= 0:
        close_p = pos['entry_price']
    pnl_pct = (close_p - pos['entry_price']) / pos['entry_price'] * 100
    market_value = pos['shares'] * close_p
    final_holdings.append({
        '代码': code, '名称': code_to_name.get(code, ''),
        '指数归属': get_index_tag(code),
        '池内': is_in_pool(code),
        '持仓数量': pos['shares'],
        '买入价格': round(pos['entry_price'], 4),
        '当前价格': round(close_p, 4),
        '盈亏%': round(pnl_pct, 2),
        '市值': round(market_value, 2),
        '入池天数': last_di - pos['entry_di'],
        '买入日期': pos['entry_date'],
    })

# 统计
final_nav = daily_nav[-1]['净值'] if daily_nav else 1.0
total_return = (final_nav - 1) * 100
n_days = len(daily_nav)
ann_return = ((final_nav ** (252 / max(1, n_days))) - 1) * 100 if final_nav > 0 else -100.0
navs = [d['净值'] for d in daily_nav]
peak = navs[0] if navs else 1.0
max_dd = 0.0
for nv in navs:
    if nv > peak:
        peak = nv
    dd = (peak - nv) / peak
    if dd > max_dd:
        max_dd = dd
sell_trades = [t for t in trades if t['买卖方向'] == '卖出']
wins = [t for t in sell_trades if t['盈亏金额'] > 0]
win_rate = len(wins) / max(1, len(sell_trades)) * 100
year_nav = {}
for d in daily_nav:
    y = d['日期'][:4]
    if y not in year_nav:
        year_nav[y] = {'start': d['净值'], 'end': d['净值']}
    else:
        year_nav[y]['end'] = d['净值']
year_returns = {}
prev_end = 1.0
for y in sorted(year_nav.keys()):
    yr = (year_nav[y]['end'] / prev_end - 1) * 100
    year_returns[y] = round(yr, 2)
    prev_end = year_nav[y]['end']
avg_hold = sum(t['持仓天数'] for t in sell_trades) / max(1, len(sell_trades))

print(f"\n{'=' * 60}")
print(f"V46_v62_v20_final (量比加分 + T+1开盘)")
print(f"最终净值: {final_nav:.4f}  ({total_return:+.2f}%)")
print(f"年化收益: {ann_return:+.2f}%  最大回撤: {max_dd * 100:.2f}%")
print(f"总交易: {len(trades)} (卖出 {len(sell_trades)}), 胜率: {win_rate:.1f}%")
print(f"平均持仓天数: {avg_hold:.1f}")
print(f"涨停跳过: {ytb_skip_count}")
tc = total_commission + total_stamp_tax
print(f"交易成本: 佣金={total_commission:,.0f} 印花税={total_stamp_tax:,.0f} 合计={tc:,.0f}")
print(f"卖出原因: {dict(sell_reason_counter)}")
print(f"量比加分: +10%={vol_bonus_count.get(0.10,0)} +20%={vol_bonus_count.get(0.20,0)} +35%={vol_bonus_count.get(0.35,0)}")

# 写出Excel (交易记录按日期降序)
now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
out_file = f'/workspace/模拟盘_V46_v62_v20_final_{now_str}.xlsx'
print(f"\n写出 Excel: {out_file}", flush=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_sheet(ws, headers, data_rows, col_widths=None):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    for ri, row in enumerate(data_rows, 2):
        for ci, h in enumerate(headers, 1):
            v = row.get(h, '')
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'


ws1 = wb.create_sheet('当前持仓')
h1 = ['代码', '名称', '指数归属', '池内', '持仓数量', '买入价格', '当前价格', '盈亏%', '市值', '入池天数', '买入日期']
write_sheet(ws1, h1, final_holdings, col_widths=[12, 14, 10, 8, 10, 12, 12, 10, 14, 10, 12])

# 交易记录: 按日期降序排列 (最新在前)
ws2 = wb.create_sheet('交易记录')
trades_sorted = sorted(trades, key=lambda x: x['日期'], reverse=True)
h2 = ['ID', '日期', '代码', '名称', '买卖方向', '价格', '数量', '金额',
      '盈亏%', '盈亏金额', '持仓天数', '卖出原因', '信号日期', '买入日期', '买入价格']
write_sheet(ws2, h2, trades_sorted, col_widths=[6, 12, 12, 14, 10, 12, 10, 14, 10, 14, 10, 16, 12, 12, 12])

ws3 = wb.create_sheet('净值曲线')
h3 = ['日期', '净值', '持仓数量', '现金', '持仓市值']
write_sheet(ws3, h3, daily_nav, col_widths=[12, 12, 12, 16, 16])

ws4 = wb.create_sheet('每日备选池Top50')
h4 = ['日期', '执行日', '排名', '代码', '名称', '指数归属', '池内', '得分', '金叉天数', 'DIF%', 'BAR%', 'vol10', '量比', '量比加分', '涨停']
write_sheet(ws4, h4, daily_top50, col_widths=[12, 12, 8, 12, 14, 10, 8, 12, 10, 10, 10, 10, 8, 10, 8])

ws5 = wb.create_sheet('最新完整备选池')
write_sheet(ws5, h4, last_full_pool, col_widths=[12, 12, 8, 12, 14, 10, 8, 12, 10, 10, 10, 10, 8, 10, 8])

# 每日排名上升最快Top10 (按日期降序)
ws_r = wb.create_sheet('每日涨幅最快Top10')
hr = ['日期', '代码', '名称', '指数归属', '当日排名', '昨日排名', '排名变化', '得分', '量比']
daily_rank_up_sorted = sorted(daily_rank_up, key=lambda x: x['日期'], reverse=True)
write_sheet(ws_r, hr, daily_rank_up_sorted, col_widths=[12, 12, 14, 10, 12, 12, 12, 12, 10])

ws6 = wb.create_sheet('统计分析')
stats = [
    {'指标': '总收益%', '数值': round(total_return, 2)},
    {'指标': '年化收益%', '数值': round(ann_return, 2)},
    {'指标': '最大回撤%', '数值': round(max_dd * 100, 2)},
    {'指标': '胜率%', '数值': round(win_rate, 2)},
    {'指标': '总交易数', '数值': len(trades)},
    {'指标': '卖出交易数', '数值': len(sell_trades)},
    {'指标': '平均持仓天数', '数值': round(avg_hold, 1)},
    {'指标': '最终净值', '数值': round(final_nav, 4)},
    {'指标': '涨停跳过次数', '数值': ytb_skip_count},
    {'指标': '交易成本合计', '数值': round(tc, 2)},
    {'指标': '回测起始日', '数值': sim_dates[0]},
    {'指标': '回测结束日', '数值': sim_dates[-1]},
    {'指标': '回测天数', '数值': len(sim_dates)},
    {'指标': '量比+10%次数', '数值': vol_bonus_count.get(0.10, 0)},
    {'指标': '量比+20%次数', '数值': vol_bonus_count.get(0.20, 0)},
    {'指标': '量比+35%次数', '数值': vol_bonus_count.get(0.35, 0)},
]
for y, r in year_returns.items():
    stats.append({'指标': f'{y}年收益%', '数值': r})
for reason, cnt in sorted(sell_reason_counter.items(), key=lambda x: -x[1]):
    stats.append({'指标': f'卖出原因:{reason}', '数值': cnt})
write_sheet(ws6, ['指标', '数值'], stats, col_widths=[28, 18])

wb.save(out_file)
print(f"Excel 已保存: {out_file}")
print(f"  Sheet1 当前持仓: {len(final_holdings)} 行")
print(f"  Sheet2 交易记录: {len(trades_sorted)} 行 (按日期降序)")
print(f"  Sheet3 净值曲线: {len(daily_nav)} 行")
print(f"  Sheet4 每日Top50: {len(daily_top50)} 行")
print(f"  Sheet5 最新备选池: {len(last_full_pool)} 行")
print(f"  Sheet6 每日涨幅最快Top10: {len(daily_rank_up_sorted)} 行 (按日期降序)")
print(f"  Sheet7 统计分析: {len(stats)} 行")
